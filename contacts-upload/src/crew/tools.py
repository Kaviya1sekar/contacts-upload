from langchain.tools import tool
import mysql.connector.pooling
import os
import fitz
import openpyxl

class ContactsUploadTools():
    @tool("File Read Tool")
    def file_read_tool(path:str):
      """
        This tool is useful to read the content of a file.
        The input to this tool should be a string representing the path to the file.
      """
      with open(path, 'rb') as file:
        content = file.read()
        content_type=path.split(".")[-1]
      text = []
      if content_type == 'pdf':
        with fitz.open(stream=content, filetype="pdf") as doc:
            for page in doc:
                text.append(page.get_text())
        return text
      elif content_type == 'xlsx':
        dataframe = openpyxl.load_workbook(path)
        dataframe_1 = dataframe.active
        for row in range(0, dataframe_1.max_row):
          for col in dataframe_1.iter_cols(1, dataframe_1.max_column):
            text.append(col[row].value)
        return text
      else:
        return content
   
    @tool("SQL Query Executor")
    def query_executor(query:str):
      """
        This tool is useful to execute an SQL query.
        The input to this tool should be a query string to be executed in SQL.
      """
      print("\nConnecting to the database")
      try:
        pool_size = 3
        db_config = {
            "host": os.environ.get("AUTOMATION_HOST"),
            "user": os.environ.get("AUTOMATION_USER"),
            "password": os.environ.get("AUTOMATION_PASSWORD"),
            "database": os.environ.get("AUTOMATION_DATABASE"),
        }
        pool = mysql.connector.pooling.MySQLConnectionPool(
              pool_name="mypool",
              pool_size=pool_size,
              pool_reset_session=True,
              **db_config)
        conn = pool.get_connection()
        curr=conn.cursor()
        print(f"\nExecuting query: {query}")
        curr.execute(query)
        conn.commit()
        print(f"\nAffected rows: {curr.rowcount}")
        conn.close()
        return "Query executed successfully"
      except Exception as e:
        return str(e)
